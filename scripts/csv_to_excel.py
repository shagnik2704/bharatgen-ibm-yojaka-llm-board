"""
CSV → Excel Converter
=====================
Reads CSV and produces a beautifully formatted XLSX file.

Usage:
    python3 scripts/csv_to_excel.py
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_CSV = Path(__file__).parent.parent / "output" / "full_questions_revised.csv"
OUTPUT_XLSX = Path(__file__).parent.parent / "output" / "full_questions_revised.xlsx"

# --- Color Palette ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ROW_EVEN_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ROW_ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Desired column widths (approximate character widths)
COLUMN_WIDTHS = {
    "course": 20,
    "block": 25,
    "bloom_level": 12,
    "citation_mode": 12,
    "question_type": 14,
    "model_id": 18,
    "language": 9,
    "question_number": 8,
    "question": 60,
    "answer": 60,
    "rubric_answer": 80,
    "rubric": 90,
    "time_taken_s": 12,
    "status": 10,
    "valid_qa": 22,
    "comments": 50,
}


def main():
    if not INPUT_CSV.exists():
        print(f"❌ CSV not found: {INPUT_CSV}")
        return

    # --- Read CSV, filter out failed rows ---
    with open(INPUT_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        all_rows = list(reader)

    # Override fieldnames to only export the requested columns
    fieldnames = [
        "course",
        "block",
        "bloom_level",
        "question",
        "answer",
        "rubric_answer",
        "rubric",
        "valid_qa",
        "comments",
    ]

    success_rows = sorted(
        [r for r in all_rows if r.get("status") == "success"],
        key=lambda r: (r.get("course", ""), r.get("block", ""), r.get("bloom_level", ""))
    )
    failed_count = len(all_rows) - len(success_rows)

    print(f"📖 Read {len(all_rows)} rows from CSV")
    if failed_count:
        print(f"   ⚠ Filtered out {failed_count} failed rows")
    print(f"   ✓ {len(success_rows)} successful rows to export")

    # --- Create workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Questions"

    # --- Friendly header names ---
    friendly_headers = {
        "course": "Course",
        "block": "Block",
        "bloom_level": "Bloom Level",
        "citation_mode": "Citation",
        "question_type": "Question Type",
        "model_id": "Model",
        "language": "Language",
        "question_number": "Q#",
        "question": "Question",
        "answer": "Answer",
        "rubric_answer": "Rubric Answer",
        "rubric": "Rubric (Marks & Key Points)",
        "time_taken_s": "Time (s)",
        "status": "Status",
        "valid_qa": "Is this a valid Question and Answer pair? (Yes or No)",
        "comments": "Comments for your decision",
    }

    # --- Write header row ---
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = friendly_headers.get(field, field)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # --- Write data rows ---
    for row_idx, row_data in enumerate(success_rows, 2):
        fill = ROW_EVEN_FILL if row_idx % 2 == 0 else ROW_ODD_FILL

        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(field, "")

            # Convert numeric fields
            if field == "time_taken_s" and value:
                try:
                    value = float(value)
                except ValueError:
                    pass
            elif field == "question_number" and value:
                try:
                    value = int(value)
                except ValueError:
                    pass

            cell.value = value
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Text wrapping for long columns
            if field in ("question", "answer", "rubric"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", horizontal="center")

    # --- Set column widths ---
    for col_idx, field in enumerate(fieldnames, 1):
        col_letter = get_column_letter(col_idx)
        width = COLUMN_WIDTHS.get(field, 15)
        ws.column_dimensions[col_letter].width = width

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions

    # --- Set row heights for content rows ---
    for row_idx in range(2, len(success_rows) + 2):
        # Taller rows for Level 2 (rubric content), shorter for Level 1
        row_data = success_rows[row_idx - 2]
        if row_data.get("rubric"):
            ws.row_dimensions[row_idx].height = 180
        else:
            ws.row_dimensions[row_idx].height = 80

    # --- Save ---
    wb.save(OUTPUT_XLSX)
    print(f"\n✅ Excel saved to: {OUTPUT_XLSX}")
    print(f"   {len(success_rows)} questions across {len(fieldnames)} columns")


if __name__ == "__main__":
    main()
