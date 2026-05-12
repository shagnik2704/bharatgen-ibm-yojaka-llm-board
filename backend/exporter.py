import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Color Palette (Matching csv_to_excel.py) ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ROW_EVEN_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ROW_ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMN_WIDTHS = {
    "course": 20,
    "block": 25,
    "bloom_level": 12,
    "question": 60,
    "answer": 60,
    "rubric": 90,
    "valid_qa": 22,
    "comments": 50,
}

FRIENDLY_HEADERS = {
    "course": "Course",
    "block": "Block",
    "bloom_level": "Bloom Level",
    "question": "Question",
    "answer": "Answer",
    "rubric": "Rubric (Marks & Key Points)",
    "valid_qa": "Is this a valid Question and Answer pair? (Yes or No)",
    "comments": "Comments for your decision",
}

FIELDNAMES = [
    "course", "block", "bloom_level", "question", "answer",
    "rubric", "valid_qa", "comments"
]

def format_rubric_text(rubric_data):
    if not rubric_data: return ""
    parts = []
    
    # Marks
    marks = rubric_data.get("marks", [])
    if marks and isinstance(marks, list):
        marks_lines = []
        for m in marks:
            if isinstance(m, dict):
                criterion = m.get("criterion", m.get("label", m.get("name", "")))
                mark_val = m.get("marks", m.get("score", m.get("value", "")))
                marks_lines.append(f"{criterion} - {mark_val} marks")
        if marks_lines:
            parts.append("[Marks]\n" + "\n".join(marks_lines))
            
    # Key Points
    key_points = rubric_data.get("key_points", [])
    if key_points and isinstance(key_points, list):
        kp_lines = [f"{i}. {kp}" for i, kp in enumerate(key_points, 1)]
        parts.append("[Key Points]\n" + "\n".join(kp_lines))
        
    return "\n\n".join(parts)

def create_session_xlsx(questions):
    """
    Takes a list of normalized question dicts and returns a BytesIO XLSX stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Questions"

    # --- Write Header Row ---
    for col_idx, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = FRIENDLY_HEADERS.get(field, field)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # --- Write Data Rows ---
    for row_idx, q in enumerate(questions, 2):
        fill = ROW_EVEN_FILL if row_idx % 2 == 0 else ROW_ODD_FILL
        
        # Mapping DB fields to Excel fields
        row_data = {
            "course": q.get("subject"),
            "block": q.get("chapter"),
            "bloom_level": q.get("depth") or q.get("bloom_level") or "",
            "question": q.get("question"),
            "answer": q.get("answer"),
            "rubric": format_rubric_text(q.get("rubric")),
            "valid_qa": "",
            "comments": ""
        }

        for col_idx, field in enumerate(FIELDNAMES, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(field, "")
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            
            # Text wrapping and alignment
            if field in ("question", "answer", "rubric", "comments", "rubric_answer"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", horizontal="center")

    # --- Set Column Widths ---
    for col_idx, field in enumerate(FIELDNAMES, 1):
        col_letter = get_column_letter(col_idx)
        width = COLUMN_WIDTHS.get(field, 15)
        ws.column_dimensions[col_letter].width = width

    # --- Freeze Header Row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions

    # --- Set Row Heights ---
    for r_idx in range(2, len(questions) + 2):
        q_data = questions[r_idx - 2]
        if q_data.get("rubric"):
            ws.row_dimensions[r_idx].height = 180
        else:
            ws.row_dimensions[r_idx].height = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
